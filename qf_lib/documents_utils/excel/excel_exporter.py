#     Copyright 2016-present CERN – European Organization for Nuclear Research
#
#     Licensed under the Apache License, Version 2.0 (the "License");
#     you may not use this file except in compliance with the License.
#     You may obtain a copy of the License at
#
#         http://www.apache.org/licenses/LICENSE-2.0
#
#     Unless required by applicable law or agreed to in writing, software
#     distributed under the License is distributed on an "AS IS" BASIS,
#     WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#     See the License for the specific language governing permissions and
#     limitations under the License.

from datetime import datetime
from os import makedirs, path, remove
from os.path import exists, isfile, join, dirname
from typing import Any, Union, Optional

import numpy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from pandas import Series, DataFrame

from qf_lib.common.tickers.tickers import Ticker
from qf_lib.common.utils.numberutils.is_finite_number import is_finite_number
from qf_lib.containers.dataframe.qf_dataframe import QFDataFrame
from qf_lib.containers.series.qf_series import QFSeries
from qf_lib.documents_utils.excel.helpers import row_and_column
from qf_lib.documents_utils.excel.write_mode import WriteMode
from qf_lib.settings import Settings
from qf_lib.starting_dir import get_starting_dir_abs_path


class ExcelExporter:

    def __init__(self, settings: Settings):
        self.settings = settings

    def export_container(
            self, container: Union[QFSeries, QFDataFrame], file_path: str,
            write_mode: WriteMode = WriteMode.CREATE_IF_DOESNT_EXIST, starting_cell: str = 'A1', sheet_name: str = None,
            include_index: bool = True, include_column_names: bool = False, remove_old_file=False) -> Union[bytes, str]:
        """
        Exports the container (QFSeries, QFDataFrame) to the excel file.
        Returns the absolute file path of the exported file.

        Parameters
        ----------
        container
            container with data to be exported
        file_path
            path (relative to the output root directory) to the file to which data should be exported
        write_mode
            mode in which the file should be opened; default: WriteMode.CREATE_IF_DOESNT_EXIST
        starting_cell
            the address of the cell which should be the top left corner of the exporter container
            default: 'A1'
        sheet_name
            the name of the sheet to which the container should be exported. If a sheet of this name doesn't exist
            it will be created. If it does: it will be edited (but not cleared). If no sheet_name is specified,
            then the currently active one will be picked
        include_index
            determines whether the index should be written together with the data.
        include_column_names
            determines whether the column names should be written together with the data. For series containers the
            column names are always "Index" and "Values".
        remove_old_file
            if true it first deletes the old file before creating new
        """
        starting_row, starting_column = row_and_column(starting_cell)

        file_path = join(get_starting_dir_abs_path(), self.settings.output_directory, file_path)
        # Make sure an old version of this file is removed.
        if remove_old_file and path.exists(file_path):
            remove(file_path)

        work_book = self.get_workbook(file_path, write_mode)
        work_sheet = self.get_worksheet(work_book, sheet_name)

        self.write_to_worksheet(
            container, work_sheet, starting_row, starting_column, include_index, include_column_names)
        work_book.save(file_path)
        return file_path

    def apply_font_style_to_area(self, file_path: str, specified_area: str, write_mode: WriteMode = WriteMode.CREATE_IF_DOESNT_EXIST, sheet_name: str = None, **font_setting):
        """
        Apply a font style to a certain area in excel

        Parameters
        ----------
        file_path
            path to the xlsl file where the cell should be written in
        specified_area: str
            the specified area where the formatting should take place in the following format: 'A1:Z12'
        write_mode
            mode in which the file should be opened
        sheet_name: str
            the name of the sheet where the cell should be written. If a sheet of this name doesn't exist
            it will be created. If it does: it will be edited (but not cleared). If no sheet_name is specified,
            then the currently active one will be picked
        font_setting
            additional font settings for the cell fonts
        """

        work_book = self.get_workbook(file_path, write_mode)
        work_sheet = self.get_worksheet(work_book, sheet_name)

        for cell in work_sheet[specified_area][0]:
            cell.font = Font(**font_setting)

        work_book.save(file_path)

    def write_cell(self, file_path: str, cell_reference: str, value: Any,
                   write_mode: WriteMode = WriteMode.CREATE_IF_DOESNT_EXIST, sheet_name: str = None):
        """
        Writes a value into the specified ``cell_reference``.

        Parameters
        ----------
        file_path
            path to the xlsl file where the cell should be written in
        cell_reference
            the address of the cell where the value should be placed, for example 'C10'
        value
            the value to write in the cell
        write_mode
            mode in which the file should be opened
        sheet_name
            the name of the sheet where the cell should be written. If a sheet of this name doesn't exist
            it will be created. If it does: it will be edited (but not cleared). If no sheet_name is specified,
            then the currently active one will be picked
        """
        row, column = row_and_column(cell_reference)
        work_book = self.get_workbook(file_path, write_mode)
        work_sheet = self.get_worksheet(work_book, sheet_name)

        self.write_to_worksheet(value, work_sheet, row, column, include_index=False, include_column_names=False)
        work_book.save(file_path)

    def get_workbook(self, file_path: str, write_mode: WriteMode) -> Workbook:
        """
        Takes a path to the file (creates it if necessary), opens the file and retrieves a Workbook object from it.
        """
        work_book = None

        if write_mode == WriteMode.CREATE_IF_DOESNT_EXIST:
            if exists(file_path) and isfile(file_path):
                write_mode = WriteMode.OPEN_EXISTING
            else:
                write_mode = WriteMode.CREATE
                dir_path = dirname(file_path)
                makedirs(dir_path, exist_ok=True)

        if write_mode == WriteMode.CREATE:
            assert not exists(file_path) or not isfile(file_path)
            work_book = Workbook()
        elif write_mode == WriteMode.OPEN_EXISTING:
            assert exists(file_path) and isfile(file_path)
            work_book = load_workbook(file_path)

        return work_book

    def get_worksheet(self, work_book: Workbook, sheet_name: str = None) -> Worksheet:
        """
        Gets a worksheet of given name from a provided workbook. If :sheet_name is None, then the active sheet
        from the workbook is returned.
        """
        if sheet_name is None:
            work_sheet = work_book.active
        else:
            work_sheet = self._get_or_create_worksheet(work_book, sheet_name)

        return work_sheet

    def write_to_worksheet(self, exported_value: Any, work_sheet: Worksheet, starting_row: int, starting_column: int,
                           include_index: bool, include_column_names: bool):
        """
        Exports a given value to Excel worksheet. If the :exported_value is a series or dataframe, then the
        :starting_row and :starting_column correspond to the top left corner of the container's values
        in the worksheet. If the :exported_value isn't a series nor dataframe, then the :include_index
        and :include_column_names parameters should be False.
        """
        if isinstance(exported_value, Series):
            self._write_series_to_worksheet(exported_value, work_sheet, starting_row, starting_column,
                                            include_index, exported_value.name if include_column_names else None)
        elif isinstance(exported_value, DataFrame):
            self._write_dataframe_to_worksheet(exported_value, work_sheet, starting_row, starting_column,
                                               include_index, include_column_names)
        else:
            assert not include_index and not include_column_names
            work_sheet.cell(row=starting_row, column=starting_column, value=self._to_supported_type(exported_value))

    def _get_or_create_worksheet(self, work_book, sheet_name):
        if sheet_name in work_book:
            work_sheet = work_book[sheet_name]
        else:
            work_sheet = work_book.create_sheet(sheet_name)

        return work_sheet

    def _write_series_to_worksheet(self, series, work_sheet, starting_row: int, starting_column: int,
                                   include_index: bool, column_name: Optional[str]):
        column = starting_column
        if include_index:
            self._export_index(series, starting_column, starting_row, work_sheet, column_name is not None)
            column += 1

        row = starting_row
        if column_name is not None:
            work_sheet.cell(row=row, column=column, value=self._to_supported_type(column_name))
            row += 1

        for date, value in series.items():
            work_sheet.cell(row=row, column=column, value=self._to_supported_type(value))
            row += 1

    def _write_dataframe_to_worksheet(self, dataframe, work_sheet, starting_row: int, starting_column: int,
                                      include_index: bool, include_column_names: bool):
        if include_index:
            self._export_index(dataframe, starting_column, starting_row, work_sheet, include_column_names)
            column = starting_column + 1
        else:
            column = starting_column

        for series_name, series in dataframe.items():
            self._write_series_to_worksheet(series, work_sheet, starting_row, column, include_index=False,
                                            column_name=series_name if include_column_names else None)
            column += 1

    def _export_index(self, container, starting_column, starting_row, work_sheet, include_column_names):
        row = starting_row
        if include_column_names:
            work_sheet.cell(row=row, column=starting_column, value="Index")
            row += 1

        for date in container.index:
            work_sheet.cell(row=row, column=starting_column, value=self._to_supported_type(date))
            row += 1

    def _to_supported_type(self, value):
        if isinstance(value, (numpy.int64, numpy.int32)):
            return int(value)
        elif isinstance(value, str) or is_finite_number(value) or isinstance(value, datetime):
            return value
        elif isinstance(value, Ticker):
            return value.as_string()
        else:
            return str(value)
